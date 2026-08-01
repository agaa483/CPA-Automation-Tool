"""Excel upload + audit endpoint.

CPA uploads a bank export spreadsheet, we run the same categorization pipeline
as the CLI's `audit-excel` command, return the annotated file.
"""
import io
import tempfile
from datetime import datetime
from pathlib import Path

import openpyxl
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse

from src import auditor
from backend.deps import require_client_ownership

router = APIRouter(
    prefix="/clients/{client_id}/excel",
    tags=["excel"],
    dependencies=[Depends(require_client_ownership)],
)


EXCEL_UNCATEGORIZED_PREFIX = "Uncategorized"


def _detect_header_row(ws) -> int:
    for r in range(1, min(11, ws.max_row + 1)):
        vals = [(str(c.value).strip().lower() if c.value else "") for c in ws[r]]
        if "date" in vals:
            return r
    raise HTTPException(status_code=400, detail="Couldn't find a header row with 'Date'")


def _find_col(header_row: list, aliases: list[str], required: bool = True) -> int | None:
    lookup = {(str(h).strip().lower() if h else ""): i + 1 for i, h in enumerate(header_row)}
    for a in aliases:
        idx = lookup.get(a.strip().lower())
        if idx is not None:
            return idx
    if required:
        raise HTTPException(
            status_code=400,
            detail=f"Missing column. Looked for any of: {aliases}",
        )
    return None


@router.post("/audit")
async def audit_excel(
    client_id: int,
    file: UploadFile = File(...),
    audit_existing: bool = Form(False),
):
    """Accept an uploaded .xlsx bank export, run the audit, return the annotated file."""
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="File must be .xlsx")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    header_row_idx = _detect_header_row(ws)
    header_row = [c.value for c in ws[header_row_idx]]

    col = {
        "date": _find_col(header_row, ["Date"]),
        "desc": _find_col(header_row, ["Bank description", "Description", "DESCRIPTION", "Memo"]),
        "spent": _find_col(header_row, ["Spent", "SPENT", "Amount Out", "Debit"]),
        "received": _find_col(header_row, ["Received", "RECEIVED", "Amount In", "Credit"]),
        "from_to": _find_col(header_row, ["From/To", "Payee", "Vendor"], required=False),
        "match": _find_col(header_row, ["Match/Categorize", "Categorize or match", "Category"]),
    }

    # Append output columns
    from openpyxl.styles import Font, PatternFill
    payee_col = ws.max_column + 1
    payor_col = ws.max_column + 2
    sugg_col = ws.max_column + 3
    reason_col = ws.max_column + 4
    ws.cell(row=header_row_idx, column=payee_col, value="Suggested Payee").font = Font(bold=True)
    ws.cell(row=header_row_idx, column=payor_col, value="Suggested Payor").font = Font(bold=True)
    ws.cell(row=header_row_idx, column=sugg_col, value="Suggested Category").font = Font(bold=True)
    ws.cell(row=header_row_idx, column=reason_col, value="Reasoning").font = Font(bold=True)

    yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    green = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

    audited = suggested = errors = 0

    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        match_val = ws.cell(row=row_idx, column=col["match"]).value
        match_str = str(match_val).strip() if match_val else ""
        is_uncategorized = (not match_str) or match_str.lower().startswith(
            EXCEL_UNCATEGORIZED_PREFIX.lower()
        )

        if audit_existing:
            if is_uncategorized:
                continue
        else:
            if not is_uncategorized:
                continue

        date_val = ws.cell(row=row_idx, column=col["date"]).value
        spent = ws.cell(row=row_idx, column=col["spent"]).value
        received = ws.cell(row=row_idx, column=col["received"]).value
        desc = ws.cell(row=row_idx, column=col["desc"]).value
        from_to = ws.cell(row=row_idx, column=col["from_to"]).value if col["from_to"] else None

        amount = spent if spent is not None else received
        direction = "out" if spent is not None else "in"

        if amount is None or date_val is None:
            continue

        audited += 1

        if audit_existing:
            txn = {
                "qbo_txn_id": f"row-{row_idx}",
                "txn_type": "Bank",
                "line_num": 1,
                "txn_date": date_val.date() if isinstance(date_val, datetime) else date_val,
                "amount": float(amount),
                "vendor_raw": desc or from_to,
                "current_qbo_category": match_str,
            }
            try:
                decision = auditor.audit_transaction(client_id, txn)
            except RuntimeError as e:
                ws.cell(row=row_idx, column=sugg_col, value=f"(error) {e}")
                errors += 1
                continue

            ws.cell(row=row_idx, column=payee_col, value=decision.suggested_payee)
            ws.cell(row=row_idx, column=payor_col, value=decision.suggested_payor)
            if decision.is_correct:
                ws.cell(row=row_idx, column=sugg_col, value="(looks correct)").fill = green
            else:
                ws.cell(row=row_idx, column=sugg_col, value=decision.corrected_category).fill = yellow
                suggested += 1
            ws.cell(row=row_idx, column=reason_col, value=decision.reasoning)
        else:
            txn = {
                "txn_date": date_val.date() if isinstance(date_val, datetime) else date_val,
                "amount": float(amount),
                "vendor_raw": desc,
                "counterparty": from_to,
                "direction": direction,
            }
            try:
                decision = auditor.suggest_category(client_id, txn)
            except RuntimeError as e:
                ws.cell(row=row_idx, column=sugg_col, value=f"(error) {e}")
                errors += 1
                continue

            ws.cell(row=row_idx, column=payee_col, value=decision.suggested_payee)
            ws.cell(row=row_idx, column=payor_col, value=decision.suggested_payor)
            if decision.is_correct:
                ws.cell(row=row_idx, column=sugg_col, value="(no suggestion)").fill = yellow
            else:
                ws.cell(row=row_idx, column=sugg_col, value=decision.corrected_category).fill = green
                suggested += 1
            ws.cell(row=row_idx, column=reason_col, value=decision.reasoning)

    # Save to a tempfile and return
    out_dir = Path(tempfile.gettempdir()) / "qb-auditor-outputs"
    out_dir.mkdir(exist_ok=True)
    out_name = f"{Path(file.filename).stem}_suggestions.xlsx"
    out_path = out_dir / out_name
    wb.save(out_path)

    return FileResponse(
        path=out_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=out_name,
        headers={
            "X-Audited": str(audited),
            "X-Suggested": str(suggested),
            "X-Errors": str(errors),
        },
    )
