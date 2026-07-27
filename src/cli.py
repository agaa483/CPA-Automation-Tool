from datetime import date as date_type, datetime
from pathlib import Path

import openpyxl
import typer
from openpyxl.styles import Font, PatternFill
from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from rich.text import Text
from tabulate import tabulate

from . import auditor, db, outlook, qbo
from .config import DB_PATH

app = typer.Typer()
console = Console()


@app.command("init-db")
def init_db() -> None:
    db.init_db()
    typer.echo(f"Database initialized at {DB_PATH}")


@app.command("add-client")
def add_client(firm_name: str = typer.Option(..., "--firm-name")) -> None:
    client_id = db.add_client(firm_name)
    typer.echo(client_id)


@app.command("qbo-connect")
def qbo_connect(client_id: int = typer.Option(..., "--client-id")) -> None:
    with db.get_connection() as conn:
        row = conn.execute(
            "SELECT qbo_refresh_token FROM clients WHERE id = ?", (client_id,)
        ).fetchone()
    if row is None:
        typer.echo(f"No client with id {client_id}.", err=True)
        raise typer.Exit(code=1)
    if row["qbo_refresh_token"]:
        confirm = typer.confirm(
            f"Client {client_id} already has QBO tokens. Overwrite?"
        )
        if not confirm:
            typer.echo("Aborted.")
            raise typer.Exit()

    try:
        qbo.start_auth_flow(client_id)
    except RuntimeError as e:
        typer.echo(f"QBO connect failed: {e}", err=True)
        raise typer.Exit(code=1)


@app.command("qbo-pull")
def qbo_pull(
    client_id: int = typer.Option(..., "--client-id"),
    days_back: int = typer.Option(30, "--days-back"),
    account: list[str] = typer.Option(
        None,
        "--account",
        help="Bank/CC account name to include. Repeat for multiple. Omit to include all bank/CC accounts.",
    ),
) -> None:
    try:
        txns = qbo.fetch_recent_transactions(
            client_id, days_back=days_back, accounts_filter=account or None
        )
    except RuntimeError as e:
        typer.echo(f"QBO pull failed: {e}", err=True)
        raise typer.Exit(code=1)

    if not txns:
        typer.echo("No transactions found.")
        return

    headers = ["qbo_txn_id", "txn_type", "line_num", "txn_date", "amount", "vendor_raw", "current_qbo_category"]
    rows = [[t[h] for h in headers] for t in txns]
    typer.echo(tabulate(rows, headers=headers, tablefmt="simple"))


@app.command("outlook-connect")
def outlook_connect(client_id: int = typer.Option(..., "--client-id")) -> None:
    with db.get_connection() as conn:
        row = conn.execute(
            "SELECT outlook_refresh_token FROM clients WHERE id = ?", (client_id,)
        ).fetchone()
    if row is None:
        typer.echo(f"No client with id {client_id}.", err=True)
        raise typer.Exit(code=1)
    if row["outlook_refresh_token"]:
        confirm = typer.confirm(
            f"Client {client_id} already has Outlook tokens. Overwrite?"
        )
        if not confirm:
            typer.echo("Aborted.")
            raise typer.Exit()

    try:
        outlook.start_auth_flow(client_id)
    except RuntimeError as e:
        typer.echo(f"Outlook connect failed: {e}", err=True)
        raise typer.Exit(code=1)


@app.command("outlook-search")
def outlook_search(
    client_id: int = typer.Option(..., "--client-id"),
    amount: float = typer.Option(..., "--amount"),
    date: str = typer.Option(..., "--date", help="YYYY-MM-DD"),
    window_days: int = typer.Option(7, "--window-days"),
) -> None:
    try:
        txn_date = datetime.strptime(date, "%Y-%m-%d").date()
    except ValueError:
        typer.echo("Invalid --date. Use YYYY-MM-DD.", err=True)
        raise typer.Exit(code=1)

    try:
        emails = outlook.search_emails(
            client_id, amount=amount, txn_date=txn_date, window_days=window_days
        )
    except RuntimeError as e:
        typer.echo(f"Outlook search failed: {e}", err=True)
        raise typer.Exit(code=1)

    if not emails:
        typer.echo("No matching emails found.")
        return

    headers = ["received_at", "sender", "subject", "body_preview"]
    rows = [[e[h] for h in headers] for e in emails]
    typer.echo(tabulate(rows, headers=headers, tablefmt="simple"))


@app.command("qbo-sync-categories")
def qbo_sync_categories(client_id: int = typer.Option(..., "--client-id")) -> None:
    try:
        count = qbo.sync_categories(client_id)
    except RuntimeError as e:
        typer.echo(f"Sync failed: {e}", err=True)
        raise typer.Exit(code=1)
    typer.echo(f"Synced {count} accounts into the categories table.")


@app.command("audit")
def audit(
    client_id: int = typer.Option(..., "--client-id"),
    txn_id: str = typer.Option(..., "--txn-id", help="QBO transaction ID"),
    line_num: int = typer.Option(
        1, "--line-num", help="Line number when the txn has multiple lines"
    ),
) -> None:
    # Pull all bank-tab txns and filter to the requested one.
    try:
        all_txns = qbo.fetch_recent_transactions(client_id)
    except RuntimeError as e:
        typer.echo(f"QBO fetch failed: {e}", err=True)
        raise typer.Exit(code=1)

    matches = [
        t for t in all_txns
        if str(t["qbo_txn_id"]) == str(txn_id) and t["line_num"] == line_num
    ]
    if not matches:
        same_id = [t for t in all_txns if str(t["qbo_txn_id"]) == str(txn_id)]
        if same_id:
            typer.echo(
                f"Txn {txn_id} has {len(same_id)} line(s). Specify --line-num "
                f"(available: {sorted(t['line_num'] for t in same_id)}).",
                err=True,
            )
        else:
            typer.echo(f"No bank-tab transaction with QBO ID {txn_id}.", err=True)
        raise typer.Exit(code=1)

    txn = matches[0]

    # Run the audit.
    try:
        decision = auditor.audit_transaction(client_id, txn)
    except RuntimeError as e:
        typer.echo(f"Audit failed: {e}", err=True)
        raise typer.Exit(code=1)

    # Persist the transaction row (needed because audit_log FK references it).
    # No-op if it already exists.
    with db.get_connection() as conn:
        conn.execute(
            """
            INSERT OR IGNORE INTO transactions (
                client_id, qbo_txn_id, txn_date, amount, vendor_raw,
                current_qbo_category
            ) VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                client_id,
                str(txn["qbo_txn_id"]),
                txn["txn_date"],
                txn["amount"],
                txn["vendor_raw"],
                txn["current_qbo_category"],
            ),
        )
        row = conn.execute(
            "SELECT id FROM transactions WHERE client_id = ? AND qbo_txn_id = ?",
            (client_id, str(txn["qbo_txn_id"])),
        ).fetchone()
        transaction_db_id = row["id"]

    auditor.log_decision(
        transaction_id=transaction_db_id,
        decision=decision,
        original_category=txn["current_qbo_category"] or "",
        prompt=getattr(decision, "_prompt", ""),
        raw_response=getattr(decision, "_raw_response", ""),
        action_taken="no_change",
    )

    # Pretty-print.
    typer.echo("")
    typer.echo(f"QBO txn:      {txn['qbo_txn_id']} (line {txn['line_num']})")
    typer.echo(f"Vendor:       {txn['vendor_raw']}")
    typer.echo(f"Amount:       {txn['amount']}")
    typer.echo(f"Existing:     {txn['current_qbo_category']}")
    typer.echo("")
    typer.echo(f"is_correct:   {decision.is_correct}")
    if not decision.is_correct:
        typer.echo(f"corrected:    {decision.corrected_category}")
    typer.echo(f"reasoning:    {decision.reasoning}")
    if decision.supporting_email_ids:
        typer.echo(f"emails cited: {', '.join(decision.supporting_email_ids)}")
    typer.echo("")
    typer.echo("(action_taken=no_change — QBO not modified.)")


def _print_dry_run_panel(txn: dict, decision, original_category: str) -> None:
    body = Text()
    body.append("Vendor    ", style="dim")
    body.append(f"{txn.get('vendor_raw') or '-'}\n")
    body.append("Amount    ", style="dim")
    body.append(f"${txn['amount']:,.2f}    ")
    body.append("Date    ", style="dim")
    body.append(f"{txn['txn_date']}\n\n")

    body.append("Current   ", style="dim")
    body.append(f"{original_category or '(none)'}\n", style="red")
    body.append("Suggest   ", style="dim")
    body.append(f"{decision.corrected_category}\n", style="green")

    body.append("\n")
    body.append("Reasoning\n", style="bold dim")
    body.append(f"  {decision.reasoning}\n", style="italic")

    if decision.supporting_email_ids:
        body.append("\n")
        body.append("Evidence  ", style="dim")
        body.append(f"{len(decision.supporting_email_ids)} email(s) cited\n")

    console.print(
        Panel(
            body,
            title=f"[bold yellow]DRY-RUN[/bold yellow]  Txn {txn['qbo_txn_id']} (line {txn['line_num']})",
            title_align="left",
            border_style="yellow",
            padding=(0, 1),
        )
    )


def _insert_new_transactions(client_id: int, txns: list[dict]) -> None:
    """Upsert each pulled txn. New rows start as 'pending'; existing rows
    have their QBO-side fields refreshed but audit_status is preserved."""
    with db.get_connection() as conn:
        for t in txns:
            conn.execute(
                """
                INSERT INTO transactions (
                    client_id, qbo_txn_id, txn_date, amount, vendor_raw,
                    current_qbo_category, audit_status
                ) VALUES (?, ?, ?, ?, ?, ?, 'pending')
                ON CONFLICT(client_id, qbo_txn_id) DO UPDATE SET
                    txn_date = excluded.txn_date,
                    amount = excluded.amount,
                    vendor_raw = excluded.vendor_raw,
                    current_qbo_category = excluded.current_qbo_category
                """,
                (
                    client_id,
                    str(t["qbo_txn_id"]),
                    t["txn_date"],
                    t["amount"],
                    t["vendor_raw"],
                    t["current_qbo_category"],
                ),
            )


def _previously_applied_original(transaction_id: int) -> str | None:
    """Return the original_category from the most recent applied audit, if any.
    Used to detect 'CPA reverted our correction' so we don't re-correct."""
    with db.get_connection() as conn:
        row = conn.execute(
            """
            SELECT original_category FROM audit_log
            WHERE transaction_id = ? AND action_taken = 'applied'
            ORDER BY id DESC LIMIT 1
            """,
            (transaction_id,),
        ).fetchone()
    return row["original_category"] if row else None


def _set_audit_status(transaction_id: int, status: str) -> None:
    with db.get_connection() as conn:
        conn.execute(
            "UPDATE transactions SET audit_status = ?, last_audited_at = CURRENT_TIMESTAMP WHERE id = ?",
            (status, transaction_id),
        )


@app.command("run")
def run(
    dry_run: bool = typer.Option(
        False, "--dry-run", help="Audit + log but do not modify QBO."
    ),
    client_id: int = typer.Option(
        None, "--client-id", help="Limit to one client. Omit to run all clients."
    ),
) -> None:
    # Resolve target clients.
    with db.get_connection() as conn:
        if client_id is not None:
            rows = conn.execute(
                "SELECT id, firm_name FROM clients WHERE id = ?", (client_id,)
            ).fetchall()
        else:
            rows = conn.execute("SELECT id, firm_name FROM clients").fetchall()
    if not rows:
        typer.echo("No clients to process.", err=True)
        raise typer.Exit(code=1)

    grand_audited = 0
    grand_no_change = 0
    grand_corrected = 0
    grand_dry_run = 0
    grand_skipped = 0
    grand_errors = 0

    for client_row in rows:
        cid = client_row["id"]
        firm = client_row["firm_name"]
        console.print()
        console.rule(f"[bold cyan]Client {cid}: {firm}[/bold cyan]")

        # Refresh tokens up-front (also surfaces auth failures before doing work).
        try:
            qbo.refresh_token_if_needed(cid)
            outlook.refresh_token_if_needed(cid)
        except RuntimeError as e:
            typer.echo(f"  Token refresh failed: {e}", err=True)
            grand_errors += 1
            continue

        # Pull from QBO and upsert new transactions.
        try:
            pulled = qbo.fetch_recent_transactions(cid)
        except RuntimeError as e:
            typer.echo(f"  QBO pull failed: {e}", err=True)
            grand_errors += 1
            continue
        _insert_new_transactions(cid, pulled)
        typer.echo(f"  Pulled {len(pulled)} txns from QBO.")

        # Build a {qbo_txn_id: txn_dict} index for feeding the auditor without re-fetching.
        pulled_index = {str(t["qbo_txn_id"]): t for t in pulled}

        # Load all pending txns for this client.
        with db.get_connection() as conn:
            pending = conn.execute(
                """
                SELECT id, qbo_txn_id, current_qbo_category
                FROM transactions
                WHERE client_id = ? AND audit_status = 'pending'
                """,
                (cid,),
            ).fetchall()
        typer.echo(f"  {len(pending)} pending txn(s) to audit.")

        for prow in pending:
            txn_db_id = prow["id"]
            qbo_id = prow["qbo_txn_id"]
            grand_audited += 1

            txn = pulled_index.get(qbo_id)
            if txn is None:
                # Pending row exists locally but not in latest pull — skip safely.
                typer.echo(f"  [{qbo_id}] not in current pull, skipping.")
                grand_skipped += 1
                continue

            # Safety: if we previously corrected this txn and it's now back to
            # the original wrong category, the CPA reverted us. Don't fight them.
            prior_orig = _previously_applied_original(txn_db_id)
            if prior_orig is not None and prior_orig == (txn["current_qbo_category"] or ""):
                _set_audit_status(txn_db_id, "do_not_audit")
                typer.echo(
                    f"  [{qbo_id}] previously corrected and reverted by CPA — "
                    f"marking do_not_audit."
                )
                grand_skipped += 1
                continue

            # Run the audit.
            try:
                decision = auditor.audit_transaction(cid, txn)
            except RuntimeError as e:
                typer.echo(f"  [{qbo_id}] audit failed: {e}", err=True)
                grand_errors += 1
                continue

            original_category = txn["current_qbo_category"] or ""
            prompt = getattr(decision, "_prompt", "")
            raw = getattr(decision, "_raw_response", "")

            if decision.is_correct:
                _set_audit_status(txn_db_id, "verified")
                auditor.log_decision(
                    transaction_id=txn_db_id,
                    decision=decision,
                    original_category=original_category,
                    prompt=prompt,
                    raw_response=raw,
                    action_taken="no_change",
                )
                typer.echo(
                    f"  [{qbo_id}] OK: '{original_category}' — {decision.reasoning}"
                )
                grand_no_change += 1
                continue

            # is_correct == False
            if dry_run:
                # Leave audit_status='pending' so a later non-dry-run picks it up.
                auditor.log_decision(
                    transaction_id=txn_db_id,
                    decision=decision,
                    original_category=original_category,
                    prompt=prompt,
                    raw_response=raw,
                    action_taken="dry_run",
                )
                _print_dry_run_panel(txn, decision, original_category)
                grand_dry_run += 1
                continue

            # Live mode: actually update QBO.
            try:
                qbo.update_category(cid, qbo_id, decision.corrected_category)
            except RuntimeError as e:
                # Log the failure but keep going through the rest.
                typer.echo(f"  [{qbo_id}] QBO update failed: {e}", err=True)
                auditor.log_decision(
                    transaction_id=txn_db_id,
                    decision=decision,
                    original_category=original_category,
                    prompt=prompt,
                    raw_response=raw,
                    action_taken="dry_run",  # treat failed apply as a dry_run record
                )
                grand_errors += 1
                continue

            _set_audit_status(txn_db_id, "corrected")
            auditor.log_decision(
                transaction_id=txn_db_id,
                decision=decision,
                original_category=original_category,
                prompt=prompt,
                raw_response=raw,
                action_taken="applied",
            )
            typer.echo(
                f"  [{qbo_id}] APPLIED: '{original_category}' -> "
                f"'{decision.corrected_category}'"
            )
            grand_corrected += 1

    # Summary
    console.print()
    summary = Table(show_header=False, box=None, padding=(0, 2))
    summary.add_column(style="dim")
    summary.add_column(justify="right", style="bold")
    summary.add_row("Audited", str(grand_audited))
    summary.add_row("No change", f"[green]{grand_no_change}[/green]")
    if dry_run:
        summary.add_row("Flagged (dry-run)", f"[yellow]{grand_dry_run}[/yellow]")
    else:
        summary.add_row("Corrected", f"[green]{grand_corrected}[/green]")
    summary.add_row("Skipped", str(grand_skipped))
    summary.add_row("Errors", f"[red]{grand_errors}[/red]" if grand_errors else "0")
    console.print(Panel(summary, title="[bold]Summary[/bold]", title_align="left", border_style="cyan"))


EXCEL_UNCATEGORIZED_PREFIX = "Uncategorized"


@app.command("audit-excel")
def audit_excel(
    client_id: int = typer.Option(..., "--client-id"),
    file: str = typer.Option(..., "--file", help="Path to the bank Excel export."),
    output: str = typer.Option(
        None, "--output", help="Output path. Defaults to <file>_suggestions.xlsx."
    ),
    audit_existing: bool = typer.Option(
        False,
        "--audit-existing",
        help="Audit rows that already have a category (check correctness). "
        "Default is to only suggest for Uncategorized rows.",
    ),
) -> None:
    in_path = Path(file)
    if not in_path.exists():
        typer.echo(f"File not found: {in_path}", err=True)
        raise typer.Exit(code=1)

    out_path = Path(output) if output else in_path.with_name(in_path.stem + "_suggestions.xlsx")

    wb = openpyxl.load_workbook(in_path)
    ws = wb.active

    # Auto-detect the header row: scan first 10 rows for one that contains "Date"
    # somewhere. Bank exports often have a title row (or two) above the headers.
    header_row_idx = None
    for r in range(1, min(11, ws.max_row + 1)):
        row_values = [
            (str(c.value).strip().lower() if c.value else "") for c in ws[r]
        ]
        if "date" in row_values:
            header_row_idx = r
            break
    if header_row_idx is None:
        typer.echo(
            "Couldn't find a header row containing 'Date' in the first 10 rows.",
            err=True,
        )
        raise typer.Exit(code=1)

    header_row = [c.value for c in ws[header_row_idx]]
    header_lookup = {
        (str(h).strip().lower() if h else ""): i + 1 for i, h in enumerate(header_row)
    }

    def find_col(*aliases: str, required: bool = True) -> int | None:
        for a in aliases:
            idx = header_lookup.get(a.strip().lower())
            if idx is not None:
                return idx
        if required:
            typer.echo(
                f"Missing expected column. Looked for any of: {aliases}. "
                f"Found headers: {[h for h in header_row if h]}",
                err=True,
            )
            raise typer.Exit(code=1)
        return None

    col = {
        "date": find_col("Date"),
        "desc": find_col("Bank description", "Description", "DESCRIPTION", "Memo"),
        "spent": find_col("Spent", "SPENT", "Amount Out", "Debit"),
        "received": find_col("Received", "RECEIVED", "Amount In", "Credit"),
        "from_to": find_col("From/To", "Payee", "Vendor", required=False),
        "match": find_col("Match/Categorize", "Categorize or match", "Category"),
    }

    # Append three output columns.
    payee_col = ws.max_column + 1
    sugg_col = ws.max_column + 2
    reason_col = ws.max_column + 3
    ws.cell(row=header_row_idx, column=payee_col, value="Suggested Payee").font = Font(bold=True)
    ws.cell(row=header_row_idx, column=sugg_col, value="Suggested Category").font = Font(bold=True)
    ws.cell(row=header_row_idx, column=reason_col, value="Reasoning").font = Font(bold=True)

    yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    green = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

    total = 0
    suggested = 0
    skipped_already = 0
    skipped_empty = 0
    errors = 0

    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        match_val = ws.cell(row=row_idx, column=col["match"]).value
        match_str = str(match_val).strip() if match_val else ""
        is_uncategorized = (not match_str) or match_str.lower().startswith(
            EXCEL_UNCATEGORIZED_PREFIX.lower()
        )

        # Skip logic depends on mode:
        # - default: only process Uncategorized rows
        # - --audit-existing: process everything that has a category set
        if audit_existing:
            if is_uncategorized:
                skipped_already += 1  # nothing to audit against
                continue
        else:
            if not is_uncategorized:
                skipped_already += 1
                continue

        date_val = ws.cell(row=row_idx, column=col["date"]).value
        spent = ws.cell(row=row_idx, column=col["spent"]).value
        received = ws.cell(row=row_idx, column=col["received"]).value
        desc = ws.cell(row=row_idx, column=col["desc"]).value
        from_to = ws.cell(row=row_idx, column=col["from_to"]).value if col["from_to"] else None

        amount = spent if spent is not None else received
        direction = "out" if spent is not None else "in"

        if amount is None or date_val is None:
            skipped_empty += 1
            continue

        total += 1

        if audit_existing:
            txn = {
                "qbo_txn_id": f"row-{row_idx}",
                "txn_type": "Bank",
                "line_num": 1,
                "txn_date": date_val.date() if isinstance(date_val, datetime) else date_val,
                "amount": float(amount),
                "vendor_raw": desc or from_to,
                "current_qbo_category": match_str,
            }
            try:
                decision = auditor.audit_transaction(client_id, txn)
            except RuntimeError as e:
                ws.cell(row=row_idx, column=payee_col, value=None)
                ws.cell(row=row_idx, column=sugg_col, value="(error)")
                ws.cell(row=row_idx, column=reason_col, value=str(e))
                errors += 1
                typer.echo(f"  Row {row_idx}: error - {e}", err=True)
                continue

            if decision.is_correct:
                ws.cell(row=row_idx, column=payee_col, value=decision.suggested_payee)
                ws.cell(row=row_idx, column=sugg_col, value="(looks correct)")
                ws.cell(row=row_idx, column=reason_col, value=decision.reasoning)
                ws.cell(row=row_idx, column=sugg_col).fill = green
                typer.echo(f"  Row {row_idx}: {desc or '(no vendor)'} ${amount} OK ({match_str})")
                typer.echo(f"    reasoning: {decision.reasoning}")
            else:
                ws.cell(row=row_idx, column=payee_col, value=decision.suggested_payee)
                ws.cell(row=row_idx, column=sugg_col, value=decision.corrected_category)
                ws.cell(row=row_idx, column=reason_col, value=decision.reasoning)
                ws.cell(row=row_idx, column=sugg_col).fill = yellow
                suggested += 1
                typer.echo(
                    f"  Row {row_idx}: {desc or '(no vendor)'} ${amount} "
                    f"FLAG '{match_str}' -> '{decision.corrected_category}'"
                )
                typer.echo(f"    reasoning: {decision.reasoning}")
        else:
            txn = {
                "txn_date": date_val.date() if isinstance(date_val, datetime) else date_val,
                "amount": float(amount),
                "vendor_raw": desc,
                "counterparty": from_to,
                "direction": direction,
            }
            try:
                decision = auditor.suggest_category(client_id, txn)
            except RuntimeError as e:
                ws.cell(row=row_idx, column=payee_col, value=None)
                ws.cell(row=row_idx, column=sugg_col, value="(error)")
                ws.cell(row=row_idx, column=reason_col, value=str(e))
                errors += 1
                typer.echo(f"  Row {row_idx}: error - {e}", err=True)
                continue

            if decision.is_correct:
                ws.cell(row=row_idx, column=payee_col, value=decision.suggested_payee)
                ws.cell(row=row_idx, column=sugg_col, value="(no suggestion)")
                ws.cell(row=row_idx, column=reason_col, value=decision.reasoning)
                ws.cell(row=row_idx, column=sugg_col).fill = yellow
            else:
                ws.cell(row=row_idx, column=payee_col, value=decision.suggested_payee)
                ws.cell(row=row_idx, column=sugg_col, value=decision.corrected_category)
                ws.cell(row=row_idx, column=reason_col, value=decision.reasoning)
                ws.cell(row=row_idx, column=sugg_col).fill = green
                suggested += 1

            typer.echo(
                f"  Row {row_idx}: {desc or '(no vendor)'} ${amount} "
                f"-> {decision.corrected_category or '(no suggestion)'}"
            )

    wb.save(out_path)

    console.print()
    summary = Table(show_header=False, box=None, padding=(0, 2))
    summary.add_column(style="dim")
    summary.add_column(justify="right", style="bold")
    summary.add_row("Rows processed", str(total))
    summary.add_row("Suggested", f"[green]{suggested}[/green]")
    summary.add_row("No suggestion", f"[yellow]{total - suggested - errors}[/yellow]")
    summary.add_row("Skipped (already categorized)", str(skipped_already))
    summary.add_row("Skipped (empty row)", str(skipped_empty))
    summary.add_row("Errors", f"[red]{errors}[/red]" if errors else "0")
    console.print(Panel(summary, title="[bold]Excel Audit Summary[/bold]", title_align="left", border_style="cyan"))
    console.print(f"\nWrote: [bold]{out_path}[/bold]")


if __name__ == "__main__":
    app()
